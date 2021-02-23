import openpyxl #패키지 불러오기
import pandas as pd
import os

wb=openpyxl.Workbook()
filename="210204.xlsx"
book= openpyxl.load_workbook(filename)
sheet = book['Sheet1']
book1= openpyxl.load_workbook("card.xlsx")
sheet1=book1['Sheet1']

print(sheet1.cell(row=2,column=2).value)

sheet1.cell(row=2,column=2).value = sheet.cell(row=2,column=1).value

print(sheet1.cell(row=2,column=2).value)
#sheet1['B2']=sheet['A2']

#print()

#get_cells= sheet['A1':'AB5']
#for row in get_cells:
#    for cell in row:
#        print(cell.value)

#get_cells= sheet1['A1':'T13']
#for row in get_cells:
#    for cell in row:
#        print(cell.value)