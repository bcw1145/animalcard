import openpyxl #패키지 불러오기
import pandas as pd
import os
from  tkinter import *
from tkinter import filedialog

root = Tk()
root.filename =  filedialog.askopenfilename(initialdir = "E:/Images",title = "choose your file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
print (root.filename)

filename=root.filename
book= openpyxl.load_workbook(filename)
sheet = book['Sheet1']
book1= openpyxl.load_workbook("cardlast.xlsx")
sheet1=book1.active

if(os.path.isdir('/Volumes/EOS_DIGITAL/cardprint/animal')==False):
    os.mkdir('/Volumes/EOS_DIGITAL/cardprint/animal')


for i in range(2,sheet.max_row+1):
    print(i)
    sheet1.cell(row=2,column=2).value = sheet.cell(row=i,column=1).value
    sheet1.cell(row=3,column=5).value = sheet.cell(row=i,column=7).value
    sheet1.cell(row=3,column=11).value = sheet.cell(row=i,column=12).value
    sheet1.cell(row=4,column=5).value = sheet.cell(row=i,column=10).value
    sheet1.cell(row=5,column=5).value = sheet.cell(row=i,column=14).value
    sheet1.cell(row=5,column=11).value = sheet.cell(row=i,column=15).value
    sheet1.cell(row=6,column=5).value = sheet.cell(row=i,column=16).value
    sheet1.cell(row=7,column=11).value = sheet.cell(row=i,column=17).value
    sheet1.cell(row=9,column=5).value = sheet.cell(row=i,column=18).value
    sheet1.cell(row=8,column=11).value = sheet.cell(row=i,column=19).value
    tmp_str=sheet.cell(row=i,column=27).value
    sheet1.cell(row=12,column=1).value =tmp_str[:4]
    sheet1.cell(row=12,column=7).value =tmp_str[6:7]
    sheet1.cell(row=12,column=12).value =tmp_str[9:10]
    book1.save('/Volumes/EOS_DIGITAL/cardprint/animal'+'/card'+str(sheet.cell(row=i,column=1).value)+'.xlsx')
