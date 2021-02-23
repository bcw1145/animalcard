import openpyxl #패키지 불러오기
import pandas as pd
import os

filename="210204.xlsx"
book= openpyxl.load_workbook(filename)
sheet = book['Sheet1']
book1= openpyxl.load_workbook("cardlast.xlsx")
sheet1=book1.active

for i in range(2,sheet.max_row+1):
    print(i)
    sheet1.cell(row=2,column=2).value = sheet.cell(row=i,column=1).value
    sheet1.cell(row=3,column=5).value = sheet.cell(row=i,column=7).value
    sheet1.cell(row=3,column=11).value = sheet.cell(row=i,column=12).value
    sheet1.cell(row=4,column=5).value = sheet.cell(row=i,column=10).value
    sheet1.cell(row=5,column=5).value = sheet.cell(row=i,column=14).value
    sheet1.cell(row=5,column=11).value = sheet.cell(row=i,column=15).value
    sheet1.cell(row=6,column=5).value = sheet.cell(row=i,column=16).value
    sheet1.cell(row=7,column=11).value = sheet.cell(row=i,column=17).value
    sheet1.cell(row=9,column=5).value = sheet.cell(row=i,column=18).value
    sheet1.cell(row=8,column=11).value = sheet.cell(row=i,column=19).value
    tmp_str=sheet.cell(row=i,column=27).value
    sheet1.cell(row=12,column=1).value =tmp_str[:4]
    sheet1.cell(row=12,column=7).value =tmp_str[6:7]
    sheet1.cell(row=12,column=12).value =tmp_str[9:10]
    book1.save('card'+str(sheet.cell(row=i,column=1).value)+'.xlsx')

#sheet1['B2']=sheet['A2']

#print()

#get_cells= sheet['A1':'AB5']
#for row in get_cells:
#    for cell in row:
#        print(cell.value)

#get_cells= sheet1['A1':'T13']
#for row in get_cells:
#    for cell in row:
#        print(cell.value)